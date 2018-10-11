#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# In[18]:

"""
0905更新日志：

2017-2018年的数据不使用 第0列(A)的chatid，改用第13列的reference_id
"""
import re
import sys
import os

from openpyxl import load_workbook

#list of all files in dir
if len(sys.argv)>1:
	arr = os.listdir(sys.argv[1])
else:
	arr = os.listdir()

print (arr)

arr_excel = [b for b in arr if b[-5:]=='.xlsx']

print (arr_excel)

output_file = [b[:-5]+"_extraction_0905.txt" for b in arr_excel]

print (output_file)

for ind,file_name in enumerate(arr_excel):

	#加载工作簿
	wb = load_workbook(filename = file_name)
	#获得表单名称
	sheets = wb.sheetnames
	ws0= wb[sheets[0]]

	new_text = "Reference ID\tChat Turn\n"


	for index,row in enumerate(ws0.rows):
	    if index<2:
	        continue

	    #如果对话数据为空，跳过
	    if row[16].value is None or len(row[16].value)<50:
	        continue
	    
	    #get agent name
	    agent_name = row[3].value[3:]
	    agent_name_= agent_name[:4]
	    
	    text_data = row[16].value
	    
	    sents_ = [b for b in re.split(r'\n',text_data)]
	    sents = [b[13:] for b in sents_]
	    
	    matches = [b[:4]==agent_name_ for b in sents]
	    
	    #如果客服的名字没有出现在对话中，或者甄别失败，跳过
	    if True not in matches:
	        continue
	    
	    id_num = 0
	    flag_new_chunk = 0
	    chunk = ""
	    buff = ""
	    
	    for a,b in zip(sents[1:],matches[1:]):
	        
	        if (a.find("' disconnected (",0,40)>0):
	            if len(chunk)>0:
	                check_buff = str(row[13].value)+"-"+str(id_num)+"\t"+chunk+"\n"
	                if check_buff != buff:
	                    new_text += str(row[13].value)+"-"+str(id_num)+"\t"+chunk+"\n"
	            continue

	        #chunk = ""

	        if b is not True:
	            client_text_pos =a.find(": ",0,50)
	            client_text = a[client_text_pos+2:]
	            #print (index,client_text)
	            if flag_new_chunk == 0:
	                chunk = client_text+" "#1010更新
	                id_num += 1
	                flag_new_chunk = 1
	            else:
	                chunk += client_text + " "#1010更新
	                
	        else:
	            if flag_new_chunk == 0:
	                continue
	            else:
	                if len(chunk)>0:
	                    buff = str(row[13].value)+"-"+str(id_num)+"\t"+chunk+"\n"
	                    new_text += str(row[13].value)+"-"+str(id_num)+"\t"+chunk+"\n"
	                    flag_new_chunk = 0
	                    continue

	text_file = open(output_file[ind], "w")

	text_file.write(new_text)

	text_file.close()
